from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

from app.score.models import ScoreRecord


@dataclass
class ScoreImportResult:
    imported: int
    skipped: int
    sheets: int


_ID_CARD_SUFFIX_LENGTH = 4
_IDENTITY_FIELDS = {"student_id", "name", "id_card_suffix"}
_LONG_FORMAT_FIELDS = _IDENTITY_FIELDS | {"course", "score"}
_NON_SCORE_HEADERS = {
    "序号",
    "学号",
    "姓名",
    "性别",
    "民族",
    "生源地",
    "书院班级",
    "行政班",
    "班级",
    "名次",
    "排名",
    "总成绩排名",
    "学分特殊加权平均成绩",
    "加权平均成绩",
    "课程名称",
    "课程号",
    "学分",
    "入学年月",
    "房间号",
    "宿舍",
    "床位号",
    "床位",
}
_HEADER_ALIASES = {
    "student_id": ["student_id", "studentid", "学号", "学生学号", "学生编号", "考号"],
    "name": ["name", "姓名", "学生姓名"],
    "class_name": ["class_name", "classname", "class", "班级", "行政班", "书院班级"],
    "id_card_suffix": [
        "id_card_suffix",
        "idcardsuffix",
        "身份证后4位",
        "身份证后四位",
        "身份证号后4位",
        "身份证号码后4位",
        "证件号后4位",
        "身份证后6位",
        "身份证后六位",
        "身份证号后6位",
        "身份证号码后6位",
        "证件号后6位",
        "身份证号",
        "身份证号码",
        "证件号",
    ],
    "total_rank": [
        "total_rank",
        "totalrank",
        "rank",
        "ranking",
        "名次",
        "排名",
        "总成绩名次",
        "总成绩排名",
        "总排名",
        "专业名次",
        "专业排名",
    ],
    "weighted_average_score": [
        "weighted_average_score",
        "weightedaveragescore",
        "weightedaverage",
        "加权平均成绩",
        "加权平均分",
        "学分特殊加权平均成绩",
        "学分加权平均成绩",
        "学分加权平均分",
        "加权平均",
        "平均成绩",
        "平均分",
        "均分",
        "总评成绩",
        "总成绩",
        "gpa",
    ],
    "course": ["course", "课程", "课程名称", "课程名", "科目", "考试科目"],
    "course_code": ["course_code", "coursecode", "课程号", "课程代码", "课程编号"],
    "credit": ["credit", "credits", "学分"],
    "score": ["score", "成绩", "分数", "得分"],
}


def _normalize_header(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace("_", "")
        .replace("-", "")
        .replace("：", "")
        .replace(":", "")
    )


_HEADER_TO_FIELD = {
    _normalize_header(alias): field
    for field, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _id_card_suffix(value: Any) -> str:
    text = _cell_to_text(value)
    token = "".join(ch.upper() for ch in text if ch.isdigit() or ch.upper() == "X")
    if not token:
        return ""
    if len(token) < _ID_CARD_SUFFIX_LENGTH and token.isdigit():
        return token.zfill(_ID_CARD_SUFFIX_LENGTH)
    return token[-_ID_CARD_SUFFIX_LENGTH:]


def _score_to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _value_to_int(value: Any) -> Optional[int]:
    score = _score_to_float(value)
    if score is None:
        return None
    return int(score)


def _header_map(row: tuple[Any, ...]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for index, value in enumerate(row):
        field = _HEADER_TO_FIELD.get(_normalize_header(value))
        if field and field not in mapping.values():
            mapping[index] = field
    return mapping


def _looks_like_score_header(value: Any) -> bool:
    header = _cell_to_text(value)
    if not header:
        return False
    normalized = _normalize_header(header)
    if not normalized or normalized in {_normalize_header(item) for item in _NON_SCORE_HEADERS}:
        return False
    return any(
        marker in normalized
        for marker in (
            "成绩",
            "分数",
            "得分",
            "期中",
            "期末",
            "考试",
            "测验",
            "score",
            "exam",
            "midterm",
            "final",
        )
    )


def _wide_score_columns(row: tuple[Any, ...], mapping: Dict[int, str]) -> Dict[int, str]:
    return {
        index: _cell_to_text(value)
        for index, value in enumerate(row)
        if index not in mapping and _looks_like_score_header(value)
    }


def _record_from_row(row: tuple[Any, ...], mapping: Dict[int, str]) -> Optional[ScoreRecord]:
    values = {
        field: row[index] if index < len(row) else None
        for index, field in mapping.items()
    }

    student_id = _cell_to_text(values.get("student_id"))
    name = _cell_to_text(values.get("name"))
    id_card_suffix = _id_card_suffix(values.get("id_card_suffix"))
    class_name = _cell_to_text(values.get("class_name"))
    total_rank = _value_to_int(values.get("total_rank"))
    weighted_average_score = _score_to_float(values.get("weighted_average_score"))
    course = _cell_to_text(values.get("course"))
    course_code = _cell_to_text(values.get("course_code"))
    credit = _score_to_float(values.get("credit"))
    score = _score_to_float(values.get("score"))

    if not student_id or not name or len(id_card_suffix) != _ID_CARD_SUFFIX_LENGTH or not course or score is None:
        return None

    return ScoreRecord(
        student_id=student_id,
        name=name,
        id_card_suffix=id_card_suffix,
        class_name=class_name,
        total_rank=total_rank,
        weighted_average_score=weighted_average_score,
        course=course,
        course_code=course_code,
        credit=credit,
        score=score,
    )


def _records_from_wide_row(
    row: tuple[Any, ...],
    mapping: Dict[int, str],
    score_columns: Dict[int, str],
) -> List[ScoreRecord]:
    values = {
        field: row[index] if index < len(row) else None
        for index, field in mapping.items()
    }

    student_id = _cell_to_text(values.get("student_id"))
    name = _cell_to_text(values.get("name"))
    id_card_suffix = _id_card_suffix(values.get("id_card_suffix"))
    class_name = _cell_to_text(values.get("class_name"))
    total_rank = _value_to_int(values.get("total_rank"))
    weighted_average_score = _score_to_float(values.get("weighted_average_score"))
    if not student_id or not name or len(id_card_suffix) != _ID_CARD_SUFFIX_LENGTH:
        return []

    records: List[ScoreRecord] = []
    for course_order, (index, course) in enumerate(score_columns.items(), start=1):
        score = _score_to_float(row[index] if index < len(row) else None)
        if score is None:
            continue
        records.append(
            ScoreRecord(
                student_id=student_id,
                name=name,
                id_card_suffix=id_card_suffix,
                class_name=class_name,
                total_rank=total_rank,
                weighted_average_score=weighted_average_score,
                course=course,
                course_order=course_order,
                score=score,
            )
        )
    return records


def _maybe_reset_dimensions(sheet: Any) -> None:
    reset_dimensions = getattr(sheet, "reset_dimensions", None)
    if callable(reset_dimensions):
        reset_dimensions()


def _summary_course_columns(
    header_row: tuple[Any, ...],
    code_row: tuple[Any, ...],
    credit_row: tuple[Any, ...],
) -> Optional[List[Dict[str, Any]]]:
    course_label_index: Optional[int] = None
    for index, value in enumerate(header_row):
        if _normalize_header(value) not in {"course", "coursename", "课程", "课程名称"}:
            continue
        if _normalize_header(code_row[index] if index < len(code_row) else "") == _normalize_header("课程号"):
            course_label_index = index
            break

    if course_label_index is None:
        return None

    columns: List[Dict[str, Any]] = []
    for index in range(course_label_index + 1, len(header_row)):
        course = _cell_to_text(header_row[index])
        if not course:
            continue
        columns.append(
            {
                "index": index,
                "course": course,
                "course_code": _cell_to_text(code_row[index] if index < len(code_row) else ""),
                "credit": _score_to_float(credit_row[index] if index < len(credit_row) else None),
                "course_order": len(columns) + 1,
            }
        )
    return columns or None


def _records_from_summary_score_sheet(sheet: Any) -> Optional[tuple[List[ScoreRecord], int]]:
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    code_row = next(rows, None)
    credit_row = next(rows, None)
    if header_row is None or code_row is None or credit_row is None:
        return None

    mapping = _header_map(header_row)
    fields = set(mapping.values())
    if not {"student_id", "name"}.issubset(fields):
        return None

    course_columns = _summary_course_columns(header_row, code_row, credit_row)
    if course_columns is None:
        return None

    records: List[ScoreRecord] = []
    skipped = 0
    for row in rows:
        values = {
            field: row[index] if index < len(row) else None
            for index, field in mapping.items()
        }
        student_id = _cell_to_text(values.get("student_id"))
        name = _cell_to_text(values.get("name"))
        if not student_id or not name:
            if any(value not in (None, "") for value in row):
                skipped += 1
            continue

        id_card_suffix = _id_card_suffix(values.get("id_card_suffix"))
        class_name = _cell_to_text(values.get("class_name"))
        total_rank = _value_to_int(values.get("total_rank"))
        weighted_average_score = _score_to_float(values.get("weighted_average_score"))

        row_records: List[ScoreRecord] = []
        for column in course_columns:
            index = column["index"]
            score = _score_to_float(row[index] if index < len(row) else None)
            if score is None:
                continue
            row_records.append(
                ScoreRecord(
                    student_id=student_id,
                    name=name,
                    id_card_suffix=id_card_suffix,
                    class_name=class_name,
                    total_rank=total_rank,
                    weighted_average_score=weighted_average_score,
                    course=column["course"],
                    course_code=column["course_code"],
                    credit=column["credit"],
                    course_order=column["course_order"],
                    score=score,
                )
            )

        if row_records:
            records.extend(row_records)
        elif any(value not in (None, "") for value in row):
            skipped += 1

    return records, skipped


def load_identity_suffixes_from_xlsx(xlsx_path: str) -> Dict[tuple[str, str], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl 未安装，无法导入身份信息表") from exc

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"身份信息表不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    suffixes: Dict[tuple[str, str], str] = {}

    try:
        for sheet in workbook.worksheets:
            _maybe_reset_dimensions(sheet)
            mapping: Optional[Dict[int, str]] = None
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if mapping is None:
                    candidate = _header_map(row)
                    if {"student_id", "name", "id_card_suffix"}.issubset(set(candidate.values())):
                        mapping = candidate
                    elif row_index >= 20:
                        break
                    continue

                values = {
                    field: row[index] if index < len(row) else None
                    for index, field in mapping.items()
                }
                student_id = _cell_to_text(values.get("student_id"))
                name = _cell_to_text(values.get("name"))
                id_card_suffix = _id_card_suffix(values.get("id_card_suffix"))
                if student_id and name and len(id_card_suffix) == _ID_CARD_SUFFIX_LENGTH:
                    suffixes[(student_id, name)] = id_card_suffix
    finally:
        workbook.close()

    return suffixes


def load_score_records_from_xlsx(xlsx_path: str) -> tuple[List[ScoreRecord], int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl 未安装，无法导入 .xlsx 成绩单") from exc

    path = Path(xlsx_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 成绩单")
    if not path.exists():
        raise FileNotFoundError(f"成绩单不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    records: List[ScoreRecord] = []
    skipped = 0

    try:
        for sheet in workbook.worksheets:
            _maybe_reset_dimensions(sheet)
            summary_records = _records_from_summary_score_sheet(sheet)
            if summary_records is not None:
                sheet_records, sheet_skipped = summary_records
                records.extend(sheet_records)
                skipped += sheet_skipped
                continue

            mapping: Optional[Dict[int, str]] = None
            score_columns: Dict[int, str] = {}
            is_wide_format = False
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if mapping is None:
                    candidate = _header_map(row)
                    candidate_fields = set(candidate.values())
                    candidate_score_columns = _wide_score_columns(row, candidate)
                    if _LONG_FORMAT_FIELDS.issubset(candidate_fields):
                        mapping = candidate
                    elif _IDENTITY_FIELDS.issubset(candidate_fields) and candidate_score_columns:
                        mapping = candidate
                        score_columns = candidate_score_columns
                        is_wide_format = True
                    elif row_index >= 20:
                        break
                    continue

                row_records = (
                    _records_from_wide_row(row, mapping, score_columns)
                    if is_wide_format
                    else [record] if (record := _record_from_row(row, mapping)) is not None else []
                )
                if not row_records:
                    if any(value not in (None, "") for value in row):
                        skipped += 1
                    continue
                records.extend(row_records)
    finally:
        workbook.close()

    return records, skipped, len(workbook.worksheets)


def import_scores_from_xlsx(
    db: Session,
    xlsx_path: str,
    replace: bool = False,
    identity_xlsx_path: str | None = None,
) -> ScoreImportResult:
    records, skipped, sheets = load_score_records_from_xlsx(xlsx_path)

    suffix_lookup: Dict[tuple[str, str], str] = {}
    if identity_xlsx_path:
        suffix_lookup.update(load_identity_suffixes_from_xlsx(identity_xlsx_path))

    for student_id, name, id_card_suffix in db.query(
        ScoreRecord.student_id,
        ScoreRecord.name,
        ScoreRecord.id_card_suffix,
    ).distinct():
        suffix = _id_card_suffix(id_card_suffix)
        if student_id and name and len(suffix) == _ID_CARD_SUFFIX_LENGTH:
            suffix_lookup.setdefault((student_id.strip(), name.strip()), suffix)

    if replace:
        db.query(ScoreRecord).delete()

    for record in records:
        if len(_cell_to_text(record.id_card_suffix)) != _ID_CARD_SUFFIX_LENGTH:
            record.id_card_suffix = suffix_lookup.get((record.student_id, record.name), "")

        db.query(ScoreRecord).filter(
            ScoreRecord.student_id == record.student_id,
            ScoreRecord.name == record.name,
            ScoreRecord.course == record.course,
        ).delete(synchronize_session=False)
        db.add(record)

    db.commit()
    return ScoreImportResult(imported=len(records), skipped=skipped, sheets=sheets)
